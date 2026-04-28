"""
src/file_handler.py
Read SKUs from uploaded CSV/Excel files.
Write enriched results to CSV bytes or Excel bytes for Streamlit download buttons.
"""

import csv
import io
from pathlib import Path


def read_file(file_bytes: bytes, filename: str) -> tuple[list[dict], list[str]]:
    """
    Read an uploaded file and return (rows, column_names).
    Each row is a dict of {column: value}.
    Supports .csv, .xlsx, .xls
    """
    ext = Path(filename).suffix.lower()

    if ext == ".csv":
        return _read_csv(file_bytes)
    elif ext in (".xlsx", ".xls"):
        return _read_xlsx(file_bytes)
    else:
        raise ValueError(f"Unsupported file type: {ext}. Please upload .csv or .xlsx")


def _read_csv(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    text    = file_bytes.decode("utf-8-sig", errors="replace")
    reader  = csv.DictReader(io.StringIO(text))
    rows    = [dict(row) for row in reader]
    columns = list(reader.fieldnames or [])
    if not columns and rows:
        columns = list(rows[0].keys())
    return rows, columns


def _read_xlsx(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    import openpyxl
    wb      = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws      = wb.active
    headers = [str(cell.value or "") for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows    = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append({headers[i]: (row[i] if i < len(row) else "") for i in range(len(headers))})
    wb.close()
    return rows, headers


def to_csv_bytes(rows: list[dict], fieldnames: list[str]) -> bytes:
    """Serialise rows to CSV bytes (UTF-8 with BOM for Excel compatibility)."""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore", lineterminator="\r\n")
    writer.writeheader()
    writer.writerows(rows)
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


def to_xlsx_bytes(rows: list[dict], fieldnames: list[str]) -> bytes:
    """Serialise rows to Excel bytes with colour-coded review_flag column."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title        = "Enriched Products"
    ws.freeze_panes = "A2"

    # Header row
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for col_idx, name in enumerate(fieldnames, 1):
        cell            = ws.cell(row=1, column=col_idx, value=name)
        cell.font       = header_font
        cell.fill       = header_fill
        cell.alignment  = Alignment(wrap_text=False)

    # Data rows
    green  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    red    = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    for row_idx, row in enumerate(rows, 2):
        flag  = str(row.get("review_flag", "") or "").upper()
        fill  = yellow if "REVIEW" in flag else red if flag in ("NOT_FOUND", "BLOCKED", "ERROR") else green

        for col_idx, field in enumerate(fieldnames, 1):
            val        = row.get(field, "") or ""
            cell       = ws.cell(row=row_idx, column=col_idx, value=str(val))
            cell.fill  = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Auto column widths
    for col_idx, field in enumerate(fieldnames, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len    = max(
            len(str(field)),
            max((len(str(r.get(field, "") or "")) for r in rows), default=0)
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 55)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_fieldnames(original_columns: list[str], output_fields: list[str]) -> list[str]:
    """Merge original columns + enrichment fields, preserving order, no duplicates."""
    seen    = set()
    ordered = []
    for col in list(original_columns) + list(output_fields):
        if col not in seen:
            seen.add(col)
            ordered.append(col)
    return ordered
